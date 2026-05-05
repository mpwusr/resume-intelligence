#!/usr/bin/env python3
"""Extract resumes (PDF/DOCX/TXT) into a single-tab xlsx (CM_Parser-style schema).

Output is one row per candidate with the columns:
    Analysis Date | CV Link | First Name | Last Name | Current Role |
    Current Employer | Prior Employers | Years Experience | Cloud Skills |
    Strategic Summary | Current Tech | Historical Tech Map | System ID

Usage:
    python extract_resumes.py --input resumes/ --output output/CV_Analysis.xlsx
"""
from __future__ import annotations

import argparse
import datetime as dt
import hashlib
import os
import sys
from pathlib import Path

import pdfplumber
from anthropic import Anthropic
from docx import Document
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from pydantic import BaseModel, Field, ValidationError

load_dotenv(override=True)

MODEL = "claude-sonnet-4-6"
MAX_TOKENS = 8192

CLOUD_VOCAB = ("AWS", "Azure", "GCP", "OCI", "IBM Cloud", "Alibaba Cloud")

SYSTEM_PROMPT = f"""You analyse resumes / CVs and emit a candidate summary as a single tool call to `extract_candidate`.

Fields:
- first_name, last_name: parse from the candidate's name.
- current_role: the most recent job title.
- current_employer: the most recent employer.
- prior_employers: list of previous employers, most recent first, current employer excluded.
- years_experience: total years of professional experience as a short string like
  "12+ years", "20 years", "5+ years". Use "+" when the resume lists "over X" or
  experience extends beyond a round figure.
- cloud_skills: subset of {list(CLOUD_VOCAB)} the candidate has hands-on experience
  with. Empty list if none mentioned.
- strategic_summary: 3–4 sentence English narrative covering leadership scope (team
  size, budget, geography), industry focus, signature achievements (cost savings,
  migrations, compliance), and certifications. Concrete numbers when the resume
  states them. No fluff, no first-person.
- current_tech: technologies / platforms / frameworks the candidate uses in their
  current role. Specific products, not categories.
- historical_tech: object mapping each prior employer name to the list of
  technologies / projects associated with that role. Use the employer name exactly
  as it appears in `prior_employers`.

Rules:
- Do not invent data. Leave a list empty (not null) when the resume is silent.
- Strategic summary must be in English regardless of resume language.
- Years experience: prefer the candidate's own claim ("over 20 years" → "20+ years")
  over computing from dates.
"""


class Candidate(BaseModel):
    first_name: str
    last_name: str
    current_role: str
    current_employer: str
    prior_employers: list[str] = Field(default_factory=list)
    years_experience: str
    cloud_skills: list[str] = Field(default_factory=list)
    strategic_summary: str
    current_tech: list[str] = Field(default_factory=list)
    historical_tech: dict[str, list[str]] = Field(default_factory=dict)


EXTRACT_TOOL = {
    "name": "extract_candidate",
    "description": "Return the candidate's structured summary.",
    "input_schema": Candidate.model_json_schema(),
}


def read_pdf(path: Path) -> str:
    out = []
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            out.append(page.extract_text() or "")
    return "\n".join(out)


def read_docx(path: Path) -> str:
    doc = Document(str(path))
    return "\n".join(p.text for p in doc.paragraphs)


def read_txt(path: Path) -> str:
    return path.read_text(encoding="utf-8", errors="replace")


def ingest(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix == ".pdf":
        return read_pdf(path)
    if suffix == ".docx":
        return read_docx(path)
    if suffix == ".txt":
        return read_txt(path)
    raise ValueError(f"Unsupported file type: {suffix}")


def extract(client: Anthropic, text: str, source: str) -> Candidate:
    user_msg = f"Source file: {source}\n\nResume text:\n\n{text}"
    last_err: Exception | None = None
    messages = [{"role": "user", "content": user_msg}]
    for _ in range(2):
        resp = client.messages.create(
            model=MODEL,
            max_tokens=MAX_TOKENS,
            system=SYSTEM_PROMPT,
            tools=[EXTRACT_TOOL],
            tool_choice={"type": "tool", "name": "extract_candidate"},
            messages=messages,
        )
        tool_use = next((b for b in resp.content if b.type == "tool_use"), None)
        if tool_use is None:
            raise RuntimeError("Model did not call extract_candidate")
        try:
            return Candidate.model_validate(tool_use.input)
        except ValidationError as e:
            last_err = e
            messages = [
                {"role": "user", "content": user_msg},
                {"role": "assistant", "content": resp.content},
                {
                    "role": "user",
                    "content": [
                        {
                            "type": "tool_result",
                            "tool_use_id": tool_use.id,
                            "content": f"Validation failed: {e}. Re-emit the tool call with the schema corrected.",
                            "is_error": True,
                        }
                    ],
                },
            ]
    raise RuntimeError(f"Extraction failed after retry: {last_err}")


def system_id(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()[:16]


def cv_link_formula(path: Path) -> str:
    uri = path.resolve().as_uri()
    label = path.name.replace('"', "'")
    return f'=HYPERLINK("{uri}", "{label}")'


def historical_tech_str(mapping: dict[str, list[str]], prior: list[str]) -> str:
    parts = []
    seen = set()
    for emp in prior:
        techs = mapping.get(emp, [])
        if not techs:
            continue
        parts.append(f"{emp.upper()}: {', '.join(techs)}")
        seen.add(emp)
    for emp, techs in mapping.items():
        if emp in seen or not techs:
            continue
        parts.append(f"{emp.upper()}: {', '.join(techs)}")
    return "; ".join(parts)


def autosize(ws, max_width: int = 80) -> None:
    for col_idx, col in enumerate(ws.columns, start=1):
        longest = 0
        for cell in col:
            v = cell.value
            if v is None:
                continue
            longest = max(longest, min(len(str(v)), max_width))
        ws.column_dimensions[get_column_letter(col_idx)].width = longest + 2


HEADERS = [
    "Analysis Date",
    "CV Link",
    "First Name",
    "Last Name",
    "Current Role",
    "Current Employer",
    "Prior Employers",
    "Years Experience",
    "Cloud Skills",
    "Strategic Summary",
    "Current Tech",
    "Historical Tech Map",
    "System ID",
]


def build_workbook(rows: list[tuple[Candidate, Path]], out_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Foglio1"
    ws.append(HEADERS)
    for cand, path in rows:
        ws.append(
            [
                dt.datetime.now(),
                cv_link_formula(path),
                cand.first_name,
                cand.last_name,
                cand.current_role,
                cand.current_employer,
                ", ".join(cand.prior_employers),
                cand.years_experience,
                ", ".join(cand.cloud_skills) if cand.cloud_skills else "N/A",
                cand.strategic_summary,
                ", ".join(cand.current_tech),
                historical_tech_str(cand.historical_tech, cand.prior_employers),
                system_id(path),
            ]
        )
    autosize(ws)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", required=True, help="Directory containing resumes")
    parser.add_argument("--output", default="output/CV_Analysis.xlsx")
    args = parser.parse_args()

    if not os.environ.get("ANTHROPIC_API_KEY"):
        print("ANTHROPIC_API_KEY is not set", file=sys.stderr)
        return 2

    in_dir = Path(args.input)
    if not in_dir.is_dir():
        print(f"Input directory not found: {in_dir}", file=sys.stderr)
        return 2

    files = sorted(
        f for f in in_dir.iterdir() if f.suffix.lower() in {".pdf", ".docx", ".txt"}
    )
    if not files:
        print(f"No PDF/DOCX/TXT files in {in_dir}", file=sys.stderr)
        return 1

    client = Anthropic()
    rows: list[tuple[Candidate, Path]] = []
    for f in files:
        try:
            text = ingest(f)
            if not text.strip():
                print(f"[skip] {f.name}: empty extraction (scanned PDF?)", file=sys.stderr)
                continue
            cand = extract(client, text, f.name)
            rows.append((cand, f))
            print(f"[ok]   {f.name}: {cand.first_name} {cand.last_name}")
        except Exception as e:
            print(f"[fail] {f.name}: {e}", file=sys.stderr)

    if not rows:
        print("No resumes successfully extracted.", file=sys.stderr)
        return 1

    out = Path(args.output)
    build_workbook(rows, out)
    print(f"Wrote {out} ({len(rows)} candidates)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
